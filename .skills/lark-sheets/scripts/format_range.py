#!/usr/bin/env python3
"""
对单元格范围一次性应用样式（字体/底色/边框/对齐/数值格式/合并/条件格式）。

用法示例：
  python scripts/format_range.py file.xlsx Sheet1 A1 --end C1 --bold --bg-color FFFF00
  python scripts/format_range.py file.xlsx Sheet1 B2:D10 --number-format "0.0%" --align center --wrap
  python scripts/format_range.py file.xlsx Sheet1 A1:E20 --cond-format \
      '{"type":"color_scale","params":{"start_type":"min","start_color":"FFAA0000","end_type":"max","end_color":"FF00AA00"}}'
"""
import argparse
import json

from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Protection, Side
from openpyxl.formatting.rule import (
    CellIsRule,
    ColorScaleRule,
    DataBarRule,
    FormulaRule,
    IconSetRule,
)

from _excel_utils import emit, emit_error, load_or_create_wb, parse_range, require_sheet


def _normalize_color(c: str) -> str:
    c = c.strip().upper().lstrip("#")
    return c if len(c) == 8 else f"FF{c}"


def _build_conditional_rule(spec: dict):
    rule_type = spec.get("type")
    params = dict(spec.get("params", {}))
    if rule_type == "color_scale":
        return ColorScaleRule(**params)
    if rule_type == "data_bar":
        return DataBarRule(**params)
    if rule_type == "icon_set":
        return IconSetRule(**params)
    if rule_type == "formula":
        return FormulaRule(**params)
    if rule_type == "cell_is":
        fill_spec = params.get("fill")
        if isinstance(fill_spec, dict):
            fg = _normalize_color(fill_spec.get("fgColor", "FFC7CE"))
            params["fill"] = PatternFill(start_color=fg, end_color=fg, fill_type="solid")
        return CellIsRule(**params)
    raise ValueError(f"Unsupported conditional format type: {rule_type}")


def main():
    p = argparse.ArgumentParser()
    p.add_argument("filepath")
    p.add_argument("sheet")
    p.add_argument("start", help="起始单元格，可写成 'A1' 或 'A1:C5'")
    p.add_argument("--end", help="结束单元格（start 用 'A1:C5' 时无需）")
    # 字体
    p.add_argument("--bold", action="store_true")
    p.add_argument("--italic", action="store_true")
    p.add_argument("--underline", action="store_true")
    p.add_argument("--font-size", type=int)
    p.add_argument("--font-color")
    p.add_argument("--font-name")
    # 填充与边框
    p.add_argument("--bg-color")
    p.add_argument("--border-style", choices=["thin", "medium", "thick", "dashed", "dotted", "double"])
    p.add_argument("--border-color")
    # 数值与对齐
    p.add_argument("--number-format")
    p.add_argument("--align", choices=["left", "center", "right", "justify"])
    p.add_argument("--vertical", choices=["top", "center", "bottom"], default="center")
    p.add_argument("--wrap", action="store_true")
    # 合并
    p.add_argument("--merge", action="store_true")
    # 保护
    p.add_argument("--protection", help='JSON 串，如 \'{"locked":true,"hidden":false}\'')
    # 条件格式
    p.add_argument("--cond-format", help="条件格式 JSON 串")
    args = p.parse_args()

    try:
        sr, sc, er, ec = parse_range(args.start, args.end)
    except ValueError as e:
        emit_error(str(e))

    wb = load_or_create_wb(args.filepath)
    try:
        ws = require_sheet(wb, args.sheet)
    except ValueError as e:
        emit_error(str(e))

    font_kwargs = {"bold": args.bold, "italic": args.italic, "underline": "single" if args.underline else None}
    if args.font_size:
        font_kwargs["size"] = args.font_size
    if args.font_name:
        font_kwargs["name"] = args.font_name
    if args.font_color:
        font_kwargs["color"] = Color(rgb=_normalize_color(args.font_color))
    font = Font(**font_kwargs)

    fill = None
    if args.bg_color:
        c = _normalize_color(args.bg_color)
        fill = PatternFill(start_color=c, end_color=c, fill_type="solid")

    border = None
    if args.border_style:
        bc = _normalize_color(args.border_color or "000000")
        side = Side(style=args.border_style, color=Color(rgb=bc))
        border = Border(left=side, right=side, top=side, bottom=side)

    align = None
    if args.align or args.wrap or args.vertical:
        align = Alignment(horizontal=args.align, vertical=args.vertical, wrap_text=args.wrap)

    protect = None
    if args.protection:
        try:
            protect = Protection(**json.loads(args.protection))
        except Exception as e:
            emit_error(f"Invalid --protection JSON: {e}")

    for r in range(sr, er + 1):
        for c in range(sc, ec + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = font
            if fill is not None:
                cell.fill = fill
            if border is not None:
                cell.border = border
            if align is not None:
                cell.alignment = align
            if protect is not None:
                cell.protection = protect
            if args.number_format:
                cell.number_format = args.number_format

    range_str = f"{args.start}" if (args.end is None and ":" not in args.start) else (
        args.start if ":" in args.start else f"{args.start}:{args.end}"
    )

    if args.merge and (er > sr or ec > sc):
        ws.merge_cells(start_row=sr, start_column=sc, end_row=er, end_column=ec)

    if args.cond_format:
        try:
            spec = json.loads(args.cond_format)
            rule = _build_conditional_rule(spec)
            ws.conditional_formatting.add(range_str, rule)
        except Exception as e:
            emit_error(f"Failed to apply conditional format: {e}")

    try:
        wb.save(args.filepath)
    except Exception as e:
        emit_error(f"Failed to save workbook: {e}")

    emit({
        "status": "success",
        "range": range_str,
        "sheet": args.sheet,
        "merged": bool(args.merge and (er > sr or ec > sc)),
        "applied": {
            "font": bool(args.bold or args.italic or args.underline or args.font_size or args.font_color or args.font_name),
            "fill": fill is not None,
            "border": border is not None,
            "align": align is not None,
            "number_format": bool(args.number_format),
            "conditional_format": bool(args.cond_format),
        },
    })


if __name__ == "__main__":
    main()
